#!/usr/bin/env python3
# Me Offer · 河北投档表解析入库
# 输入：data_raw/hebei/hb_YYYY_physics.xlsx 和 hb_YYYY_history.xlsx
# 输出：data_raw/hebei/hb_YYYY_insert.sql

import os
import re
import json
import sys
import openpyxl

DATA_DIR							= os.path.join(os.path.dirname(__file__), '..', 'data_raw', 'hebei')

YEAR_SOURCES						= {
	2025: {
		'physics':					'https://tangshan.huanbohainews.com.cn/resource/2025-07/23/9914ae17-e100-497a-986d-484d3b9a1c0f.xlsx',
		'history':					'https://imgbdb3.bendibao.com/sjzbdb/edu/20257/23/2025723093850_46223.xlsx'
	},
	2024: {
		'physics':					'https://gaokzx.com/cdn (via Wayback Machine) · file.hebeea.edu.cn 2024-07',
		'history':					'https://huanbohainews.com.cn/cdn (via Wayback Machine) · file.hebeea.edu.cn 2024-07'
	},
	2023: {
		'physics':					'https://web.archive.org/web/file.hebeea.edu.cn/files/article/2023/07/',
		'history':					'https://web.archive.org/web/file.hebeea.edu.cn/files/article/2023/07/'
	},
	2022: {
		'physics':					'https://web.archive.org/web/file.hebeea.edu.cn/files/article/2022/07/',
		'history':					'https://web.archive.org/web/file.hebeea.edu.cn/files/article/2022/07/'
	},
	2021: {
		'physics':					'https://web.archive.org/web/file.hebeea.edu.cn/files/article/2021/07/',
		'history':					'https://web.archive.org/web/file.hebeea.edu.cn/files/article/2021/07/'
	}
}


def parse_school_name(raw):
	name							= raw.strip()
	city							= ''
	nature							= ''

	m_nature						= re.search(r'\[([^\]]+)\]', name)

	if m_nature:
		nature						= m_nature.group(1).strip()
		name						= re.sub(r'\[[^\]]+\]', '', name).strip()

	m_city							= re.search(r'\(([^)]+)\)', name)

	if m_city:
		city						= m_city.group(1).strip()
		name						= re.sub(r'\([^)]+\)', '', name).strip()

	return name, city, nature


def parse_one(xlsx_path, year, subject_type, source_url):
	wb								= openpyxl.load_workbook(xlsx_path, data_only=True)
	sh								= wb.active

	rows							= []
	skipped							= 0

	for r in range(6, sh.max_row + 1):
		school_code					= sh.cell(row=r, column=1).value
		school_raw					= sh.cell(row=r, column=2).value
		group_code					= sh.cell(row=r, column=3).value
		group_name					= sh.cell(row=r, column=4).value
		min_score					= sh.cell(row=r, column=5).value
		ts_lang_math				= sh.cell(row=r, column=6).value
		ts_lang_max					= sh.cell(row=r, column=7).value
		ts_foreign					= sh.cell(row=r, column=8).value
		ts_first					= sh.cell(row=r, column=9).value
		ts_second_max				= sh.cell(row=r, column=10).value
		ts_second_sub				= sh.cell(row=r, column=11).value
		ts_vol_no					= sh.cell(row=r, column=12).value

		if school_code is None or school_raw is None:
			skipped					+= 1
			continue

		school_code					= str(school_code).strip()
		group_code					= str(group_code).strip() if group_code is not None else ''
		group_name					= str(group_name).strip() if group_name is not None else ''

		if not school_code or not group_name:
			skipped					+= 1
			continue

		try:
			min_score				= int(float(min_score)) if min_score is not None else None
		except (ValueError, TypeError):
			min_score				= None

		if min_score is None:
			skipped					+= 1
			continue

		school_name, school_city, school_nature = parse_school_name(str(school_raw))

		tiebreak					= {
			'lang_math':			ts_lang_math,
			'lang_max':				ts_lang_max,
			'foreign':				ts_foreign,
			'first':				ts_first,
			'second_max':			ts_second_max,
			'second_sub':			ts_second_sub,
			'vol_no':				ts_vol_no
		}

		for k in list(tiebreak.keys()):
			v						= tiebreak[k]

			try:
				tiebreak[k]			= int(float(v)) if v is not None else None
			except (ValueError, TypeError):
				tiebreak[k]			= None

		rows.append({
			'year':					year,
			'province':				'hebei',
			'subject_type':			subject_type,
			'school_code':			school_code,
			'school_name':			school_name,
			'school_city':			school_city,
			'school_nature':		school_nature,
			'group_code':			group_code,
			'group_name':			group_name,
			'min_score':			min_score,
			'tiebreak_json':		json.dumps(tiebreak, ensure_ascii=False),
			'source_url':			source_url
		})

	print('  year', year, subject_type, ': parsed', len(rows), 'rows, skipped', skipped)
	return rows


def escape(s):
	return s.replace("'", "''") if s else ''


def main():
	all_rows						= []

	for year, sources in YEAR_SOURCES.items():
		for subject_type, url in sources.items():
			fname					= 'hb_' + str(year) + '_' + subject_type + '.xlsx'
			fpath					= os.path.join(DATA_DIR, fname)

			if not os.path.exists(fpath):
				print('MISSING:', fname)
				continue

			rows					= parse_one(fpath, year, subject_type, url)
			all_rows.extend(rows)

	print()
	print('total:', len(all_rows))

	by_year							= {}

	for r in all_rows:
		by_year.setdefault(r['year'], []).append(r)

	BATCH							= 80

	for year in sorted(by_year.keys()):
		rows						= by_year[year]
		out_sql						= os.path.join(DATA_DIR, 'hebei_' + str(year) + '_insert.sql')

		with open(out_sql, 'w', encoding='utf-8') as f:
			f.write('-- Me Offer · 河北 ' + str(year) + ' 投档表入库 · 共 ' + str(len(rows)) + ' 条\n\n')

			for i in range(0, len(rows), BATCH):
				batch				= rows[i:i+BATCH]
				values				= []

				for r in batch:
					tj				= r['tiebreak_json'].replace("'", "''")
					values.append(
						"(" + str(r['year']) + ",'hebei','" + r['subject_type'] + "','" +
						escape(r['school_code']) + "','" + escape(r['school_name']) + "','" +
						escape(r['school_city']) + "','" + escape(r['school_nature']) + "','" +
						escape(r['group_code']) + "','" + escape(r['group_name']) + "'," +
						str(r['min_score']) + ",'" + tj + "','" + escape(r['source_url']) + "')"
					)

				f.write('INSERT INTO gaokao_scores (year, province, subject_type, school_code, school_name, school_city, school_nature, group_code, group_name, min_score, tiebreak_json, source_url) VALUES\n')
				f.write(',\n'.join(values))
				f.write(';\n\n')

		print('saved:', out_sql, '(' + str(len(rows)) + ' rows)')

	stats							= {}

	for r in all_rows:
		key							= (r['year'], r['subject_type'])
		stats[key]					= stats.get(key, 0) + 1

	print()
	print('year x subject_type stats:')

	for k, v in sorted(stats.items()):
		print(' ', k, ':', v)


if __name__ == '__main__':
	main()
